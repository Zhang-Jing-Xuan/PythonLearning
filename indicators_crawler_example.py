import requests
from pyquery import PyQuery as pq#基于CSS选择器的解析库
from selenium import webdriver#自动化爬虫，与chromedriver搭配使用
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl#向excel导入数据
import pandas as pd
import time
import json
import os
import gc

code_url = 'http://quotes.money.163.com/old/#query=hy003026&DataType=HS_RANK&sort=PERCENT&order=desc&count=24&page=0'
time_column = ['2019-09-30', '2019-06-30', '2019-03-31', '2018-12-31', '2018-09-30', '2018-06-30']
sheet_column = ['名称', '2019-09-30', '2019-06-30', '2019-03-31', '2018-12-31', '2018-09-30', '2018-06-30']
sheet_name = ['总资产利润率(%)', '销售净利率(%)', '流动比率(%)', '资产负债率(%)', '总资产周转率(次)', '流动资产周转率(次)', '速动比率(%)']


def get_code(url):
    code_list = []
    chromedriver = "/usr/local/bin/chromedriver"#chromedriver存放位置
    browser = webdriver.Chrome(chromedriver)
    wait = WebDriverWait(browser, 10)#最多延迟10秒
    browser.get(url)
    # browser.add_cookie(cookie)
    for i in range(1, 6):#爬取1-5页
        print(f'Crawling Page {i}')
        try:
            if i == 1:#按股票代码排序，防止因为实时数据导致爬取相同数据
                code_sort = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'table.table-header > thead > tr > th:nth-child(2)')))
                code_sort.click()
            time.sleep(0.5)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'table.ID_table.stocks-info-table > tbody > tr')))
            html = browser.page_source
            doc = pq(html)
            table = doc('table.ID_table.stocks-info-table > tbody > tr').items()
            for tr in table:
                code_list.append({
                    'code': tr.find('td:nth-child(2) > a').text(),
                    'name': tr.find('td:nth-child(3) a.symbol').text()
                })
            # print(code_list)
            page_flip = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.ID_pages.mod-pages > a:last-child')))#点击最后一页
            page_flip.click()
        except:
            print(f'Page {i} does not exist.')
            break
    # save to json
    with open('code_list.json', 'w', encoding='utf-8') as f:
        f.write(json.dumps(code_list))
    return code_list


def parse_page(code, name):
    url = f'http://quotes.money.163.com/f10/zycwzb_{code}.html#01c01'
    response = requests.get(url)
    if response.status_code == 200:#200表示请求成功
        doc = pq(response.text)
        indicators = {
            '代码': code,#函数参数
            '名称': name,#函数参数
            '总资产利润率(%)': doc('td:contains("总资产利润率(%)")').siblings().text().split(),
            '销售净利率(%)': doc('td:contains("销售净利率(%)")').siblings().text().split(),
            '流动比率(%)': doc('td:contains("流动比率(%)")').siblings().text().split(),
            '资产负债率(%)': doc('td:contains("资产负债率(%)")').siblings().text().split(),
            '总资产周转率(次)': doc('td:contains("总资产周转率(次)")').siblings().text().split(),
            '流动资产周转率(次)': doc('td:contains("流动资产周转率(次)")').siblings().text().split(),
            '速动比率(%)': doc('td:contains("速动比率(%)")').siblings().text().split(),
        }
        print(f'success: {code}')
        print(indicators)
        return indicators
    else:
        print(f'error: {code}')
        return 0


def get_indicator(code_list):
    res = []
    for i in code_list:
        indicators = parse_page(i['code'], i['name'])#爬取单页
        if indicators:
            res.append(indicators)
    with open('indicators.json', 'w', encoding='utf-8') as f:
        f.write(json.dumps(res))


def ensure_data():
    if not os.path.exists('code_list.json'):#创建出股票代码和股票名称的.json
        get_code(code_url)
    if not os.path.exists('indicators.json'):#创建出该股票中所要的信息.json
        with open('code_list.json', 'r', encoding='utf-8') as f:
            code_list = json.load(f)#将json格式字符串转化为dict
        get_indicator(code_list)


def to_excel(data):
    file_name = 'indicators.xlsx'
    print('Writing to excel...')
    # create new sheets
    wb = openpyxl.Workbook(file_name)
    for s in sheet_name:
        ws = wb.create_sheet(s)
    wb.save(file_name)
    # preserve created sheets
    book = openpyxl.load_workbook(file_name)#读取
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # restructure and write data
    for sheet in sheet_name:
        print(f'Processing: {sheet}')
        result = {}
        for val in data:
            code = val["代码"]
            result[code] = {}#创建空字典
            result[code]["名称"] = val["名称"]
            for i in range(len(time_column)):
                try:
                    result[code][time_column[i]] = float(val[sheet][i])
                except:
                    result[code][time_column[i]] = None

        df = pd.DataFrame(result, index=pd.Series(sheet_column)).T
        df.to_excel(writer, sheet_name=sheet)
        writer.save()
    # release memory
    print('Finished.')
    del wb, ws
    gc.collect()


def main():
    ensure_data()
    with open('indicators.json', 'r', encoding='utf-8') as f:#不必调用f.close()方法
        data = json.load(f)#将json格式字符串转化为dict
        to_excel(data)#写入excel


if __name__ == '__main__':
    main()
    #parse_page(688418, '震有科技')
    # with open('code_list.json', 'r', encoding='utf-8') as f:
    #     code_list = json.load(f)
    # get_indicator(code_list)
