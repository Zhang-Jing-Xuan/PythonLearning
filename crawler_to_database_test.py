import requests
import decimal
from pyquery import PyQuery as pq
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
import pymysql
import pandas as pd
import time
import json
import os
import gc

# 爬取的网页网址
code_url = 'http://quotes.money.163.com/old/#query=hy009000&DataType=HS_RANK&sort=PERCENT&order=desc&count=24&page=0'
# 时间段
time_column = ['2020-06-30', '2020-03-31', '2019-12-31', '2019-09-30', '2019-06-30', '2019-03-31']
# 表格的字段名
sheet_column = ['名称', '2020-06-30', '2020-03-31', '2019-12-31', '2019-09-30', '2019-06-30', '2019-03-31']
# 各表格名称
sheet_name = ['总资产利润率(%)', '销售净利率(%)']
# 各数据表名称
table_name = ['总资产利润率', '销售净利率']


# 获取所有的股票编号和名称
def get_code(url):
    code_list = []
    chromedriver = "/usr/local/bin/chromedriver"#chromedriver存放位置
    browser = webdriver.Chrome(chromedriver)
    wait = WebDriverWait(browser, 10)      # 设置等待时间
    browser.get(url)      # 浏览器打开对应网址
    # browser.add_cookie(cookie)
    for i in range(1, 3):     # 爬取页数
        print(f'Crawling Page {i}')
        try:
            if i == 1:
                # 等待页面显示后获取股票序号这个字段标题。nth-child() 选择器，nth-child(2)其父元素的第二个子元素
                code_sort = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'table.table-header > thead > tr > th:nth-child(2)')))
                # 对股票进行点击排序
                code_sort.click()
            # 将程序短暂的停顿一下
            time.sleep(0.5)
            # 等待页面显示后获取对应元素的值
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'table.ID_table.stocks-info-table > tbody > tr')))
            # page_source页面源代码
            html = browser.page_source
            # 解析html
            doc = pq(html)
            # 获得指定部分的html。.items()获取所有的tr
            table = doc('table.ID_table.stocks-info-table > tbody > tr').items()
            # 遍历表格中的每一行，获得编号和名字
            for tr in table:
                code_list.append({
                    'code': tr.find('td:nth-child(2) > a').text(),
                    'name': tr.find('td:nth-child(3) a.symbol').text()
                })
            # print(code_list)
            # 获取下一页按钮的值
            page_flip = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.ID_pages.mod-pages > a:last-child')))
            # 点击跳转到下一页
            page_flip.click()
        except:
            print(f'Page {i} does not exist.')     #  出错则打印第i页没有爬取
            break
    # save to json
    with open('code_list.json', 'w', encoding='utf-8') as f:
        f.write(json.dumps(code_list))
    return code_list


# 获得每一个股票的具体信息
def parse_page(code, name):
    url = f'http://quotes.money.163.com/f10/zycwzb_{code}.html#01c01'  # 根据编码确定相应的网页
    response = requests.get(url)  # 发送GET请求，返回Response对象
    if response.status_code == 200:
        doc = pq(response.text)
        indicators = {
            '代码': code,
            '名称': name,
            # .siblings() 子组件 / contains()是选择对象中包括指定内容的对象
            '总资产利润率(%)': doc('td:contains("总资产利润率(%)")').siblings().text().split(),
            '销售净利率(%)': doc('td:contains("销售净利率(%)")').siblings().text().split(),
            '流动比率(%)': doc('td:contains("流动比率(%)")').siblings().text().split(),
            '资产负债率(%)': doc('td:contains("资产负债率(%)")').siblings().text().split(),
            '总资产周转率(次)': doc('td:contains("总资产周转率(次)")').siblings().text().split(),
            '流动资产周转率(次)': doc('td:contains("流动资产周转率(次)")').siblings().text().split(),
            '速动比率(%)': doc('td:contains("速动比率(%)")').siblings().text().split(),
        }
        print(f'success: {code}')
        print(indicators)
        return indicators
    else:
        print(f'error: {code}')
        return 0


# 将股票的具体信息写入json文件中
def get_indicator(code_list):
    res = []
    for i in code_list:
        indicators = parse_page(i['code'], i['name'])
        if indicators:
            res.append(indicators)
    with open('indicators.json', 'w', encoding='utf-8') as f:
        f.write(json.dumps(res))


# 确认文件是否存在
def ensure_data():
    if not os.path.exists('code_list.json'):
        get_code(code_url)
    if not os.path.exists('indicators.json'):
        with open('code_list.json', 'r', encoding='utf-8') as f:
            code_list = json.load(f)
        get_indicator(code_list)


# 将信息填入excel表格中
def to_excel(data):
    file_name = 'indicators.xlsx'
    print('Writing to excel...')
    # create new sheets
    wb = openpyxl.Workbook(file_name)    # 创建一个excel文件
    for s in sheet_name:
        ws = wb.create_sheet(s)    # 创建表
    wb.save(file_name)    # 保存文件
    # preserve created sheets
    book = openpyxl.load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    # dict() 函数用于创建一个字典
    # restructure and write data
    for sheet in sheet_name:
        print(f'Processing: {sheet}')
        result = {}
        for val in data:
            code = val["代码"]
            result[code] = {}
            result[code]["名称"] = val["名称"]
            for i in range(len(time_column)):
                try:
                    result[code][time_column[i]] = float(val[sheet][i])
                except:
                    result[code][time_column[i]] = None

        df = pd.DataFrame(result, index=pd.Series(sheet_column)).T
        df.to_excel(writer, sheet_name=sheet)   # 写入
        writer.save()
    # release memory
    print('Finished.')
    del wb, ws
    gc.collect()     # 清除内存


# 将数据添加到mysql数据库中
def to_mysql(data):
    '''
    # 创建数据库
    conn = pymysql.connect(
        host='localhost',
        user='root',
        password='271828',
    )
    # 创建游标
    cursor = conn.cursor()
    cursor.execute('CREATE DATABASE Aug23;')  # 使用连接对象获得一个cursor对象进行工作.这些方法包括两大类:1.执行命令,2.接收返回值
    # 如果执行上述代码没报错，那么已经成功创建了一个数据库
    conn.commit()
    conn.close()
    '''
    # 连接数据库
    db = pymysql.connect(
        host="localhost",
        user="root",
        passwd="271828",
        db='dbtest'
    )
    # 创建游标
    mycursor = db.cursor()

    # 创建数据表，设置字段属性
    for t in table_name:
        sql = """
                create table """ + str(t) + """(
                id char(10),
                name char(10),
                2020_06_30 char(10),
                2020_03_31 char(10),
                2019_12_31 char(10),
                2019_09_30 char(10), 
                2019_06_30 char(10), 
                2019_03_31 char(10))default charset = utf8
            """
        try:
            # 执行SQL语句
            mycursor.execute(sql)
        except :
            print("创建数据表%s失败" %(t))

    # 插入数据
    for i in range(len(table_name)):
        print(f'Processing: {table_name[i]}')
        s = sheet_name[i]
        for val in data:
            # 将每一条记录的信息保存到result列表中

            result = [val["代码"], val["名称"], val[s][0], val[s][1],
                 val[s][2], val[s][3],
                 val[s][4], val[s][5]]
            print(result)

            strSql = """ INSERT INTO """ + str(table_name[i]) + """ 
                        (id, name , 2020_06_30 , 2020_03_31 , 2019_12_31 ,2019_09_30 ,2019_06_30 ,2019_03_31 )
                        VALUES  (%s, %s, %s, %s, %s, %s, %s, %s)"""

            mycursor.execute(strSql, result)   # 插入记录
            result = []


    db.commit()
    db.close()

    print('Finished.')
    gc.collect()     # release memory

def main():
    ensure_data()
    with open('indicators.json', 'r', encoding='utf-8') as f:
        data = json.load(f)    # 从json文件中读取数据
        #to_excel(data)      # 将信息填入excel表格中
        to_mysql(data)      # 将信息填入mysql数据库中

if __name__ == '__main__':
    main()
    # parse_page(688418, '震有科技')
    # with open('code_list.json', 'r', encoding='utf-8') as f:
    #     code_list = json.load(f)
    # get_indicator(code_list)
    '''
    with open('indicators.json', 'r', encoding='utf-8') as f:
        data = json.load(f)    # 从json文件中读取数据
        to_mysql(data)
    '''
